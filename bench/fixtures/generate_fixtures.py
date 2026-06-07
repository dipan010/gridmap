"""Generate xlsx test fixtures for gridmap benchmark/integration testing.

Each fixture tests a specific detection pathway. Run this script to
regenerate all 14 fixtures from scratch (reproducible, no manual editing).

Usage:
    python generate_fixtures.py
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.comments import Comment


OUTPUT_DIR = Path(__file__).parent


def save(wb: openpyxl.Workbook, name: str) -> None:
    wb.save(OUTPUT_DIR / name)
    wb.close()


# ---------- 01: Adjacent right ----------

def gen_01():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Password"
    ws["B1"] = "s3cRet!99"
    save(wb, "01_adjacent_right.xlsx")


# ---------- 02: Adjacent below ----------

def gen_02():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Password"
    ws["A2"] = "MyP@ss123"
    save(wb, "02_adjacent_below.xlsx")


# ---------- 03: Inline same cell ----------

def gen_03():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "password: hunter2!"
    save(wb, "03_inline_same_cell.xlsx")


# ---------- 04: Formula concatenation ----------

def gen_04():
    wb = openpyxl.Workbook()
    ws = wb.active
    # Strategy 2: keyword in one string, value in the next
    ws["A1"] = '=CONCAT("password","S3cure#1")'
    save(wb, "04_formula_concat.xlsx")


# ---------- 05: Comment on header ----------

def gen_05():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Password"
    ws["A1"].comment = Comment("xK9$mL2p", "admin")
    save(wb, "05_comment_on_header.xlsx")


# ---------- 06: Comment inline ----------

def gen_06():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Server Config"
    ws["A1"].comment = Comment("password: abc12345", "admin")
    save(wb, "06_comment_inline.xlsx")


# ---------- 07: Split password ----------

def gen_07():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Password"
    ws["A2"] = "ab"
    ws["A3"] = "cd"
    ws["A4"] = "ef"
    save(wb, "07_split_password.xlsx")


# ---------- 08: Multi-sheet ----------

def gen_08():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Settings"
    ws1["A1"] = "Password"
    ws1["B1"] = "Sheet1Pw!"

    wb.create_sheet("Data")
    ws2 = wb["Data"]
    ws2["A1"] = "Name"
    ws2["B1"] = "Alice"

    wb.create_sheet("Admin")
    ws3 = wb["Admin"]
    ws3["A1"] = "Token"
    ws3["B1"] = "tk_9Xf$2m"

    save(wb, "08_multi_sheet.xlsx")


# ---------- 09: Hidden sheet ----------

def gen_09():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Public"
    ws1["A1"] = "Welcome"

    ws2 = wb.create_sheet("Secrets")
    ws2["A1"] = "Password"
    ws2["B1"] = "H1dd3n!x"
    ws2.sheet_state = "hidden"

    save(wb, "09_hidden_sheet.xlsx")


# ---------- 10: With context ----------

def gen_10():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Username"
    ws["B1"] = "admin"
    ws["A2"] = "Password"
    ws["B2"] = "Str0ng!Pw"
    ws["A3"] = "URL"
    ws["B3"] = "https://example.com"
    save(wb, "10_with_context.xlsx")


# ---------- 11: Noise ----------

def gen_11():
    wb = openpyxl.Workbook()
    ws = wb.active
    # Fill 100 noise cells across 10 rows x 10 cols
    for row in range(1, 11):
        for col in range(1, 11):
            ws.cell(row=row, column=col, value=f"data_{row}_{col}")
    # Place credential on row 12 (below noise) with nothing below it
    ws.cell(row=12, column=5, value="Password")
    ws.cell(row=12, column=6, value="Bu1r3d!x")
    save(wb, "11_noise.xlsx")


# ---------- 12: Multilingual ----------

def gen_12():
    wb = openpyxl.Workbook()
    ws = wb.active
    # German
    ws["A1"] = "Passwort"
    ws["B1"] = "Ge#rm4n1"
    # Spanish
    ws["A2"] = "Contraseña"
    ws["B2"] = "Sp@n1sh!"
    # Japanese
    ws["A3"] = "パスワード"
    ws["B3"] = "Jp_T0k3n"
    save(wb, "12_multilingual.xlsx")


# ---------- 13: No credentials ----------

def gen_13():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Name"
    ws["B1"] = "Alice"
    ws["A2"] = "Age"
    ws["B2"] = "30"
    ws["A3"] = "City"
    ws["B3"] = "Portland"
    ws["A4"] = "Score"
    ws["B4"] = "95"
    save(wb, "13_no_credentials.xlsx")


# ---------- 14: Merged cells ----------

def gen_14():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Password"
    ws.merge_cells("A1:A3")
    ws["B1"] = "M3rg3d!x"
    save(wb, "14_merged_cells.xlsx")


# ---------- Main ----------

GENERATORS = [
    gen_01, gen_02, gen_03, gen_04, gen_05, gen_06, gen_07,
    gen_08, gen_09, gen_10, gen_11, gen_12, gen_13, gen_14,
]


def main() -> None:
    for gen in GENERATORS:
        gen()
    print(f"Generated {len(GENERATORS)} fixture files in {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
