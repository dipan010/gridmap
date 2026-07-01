"""Tests for gridmap.api — load(), GridDoc, Relationship."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

import gridmap
from gridmap.api import GridDoc, Relationship, load


@pytest.fixture
def tmp_xlsx(tmp_path: Path):
    """Helper that writes an openpyxl Workbook to a temp file and returns the path."""

    def _write(wb: openpyxl.Workbook, name: str = "test.xlsx") -> Path:
        p = tmp_path / name
        wb.save(p)
        return p

    return _write


@pytest.fixture
def credential_xlsx(tmp_xlsx):
    """An xlsx file with a Password / s3cret!! credential."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Password"
    ws["B1"] = "s3cret!!"
    p = tmp_xlsx(wb)
    wb.close()
    return p


# ---------- load ----------


def test_load_returns_griddoc(credential_xlsx):
    doc = load(credential_xlsx)
    assert isinstance(doc, GridDoc)


def test_load_filepath_is_path(credential_xlsx):
    doc = load(credential_xlsx)
    assert isinstance(doc.filepath, Path)
    assert doc.filepath == credential_xlsx


def test_load_sheet_count(credential_xlsx):
    doc = load(credential_xlsx)
    assert doc.sheet_count == 1


def test_load_cell_count(credential_xlsx):
    doc = load(credential_xlsx)
    assert doc.cell_count == 2


def test_load_accepts_string(credential_xlsx):
    doc = load(str(credential_xlsx))
    assert isinstance(doc, GridDoc)


def test_griddoc_is_frozen(credential_xlsx):
    doc = load(credential_xlsx)
    with pytest.raises(AttributeError):
        doc.filepath = Path("/tmp/nope")


def test_load_nonexistent_raises():
    with pytest.raises(FileNotFoundError):
        load("/tmp/does_not_exist_gridmap_test.xlsx")


def test_load_unsupported_extension_raises(tmp_path):
    """An unsupported file extension should be rejected."""
    p = tmp_path / "test.docx"
    p.write_text("not a spreadsheet")
    with pytest.raises(ValueError, match=r"Unsupported file extension"):
        load(p)


def test_load_wrong_content_raises(tmp_path):
    """A .xlsx file that is actually CSV should be rejected by magic-byte check."""
    p = tmp_path / "fake.xlsx"
    p.write_text("a,b,c\n1,2,3")
    with pytest.raises(ValueError, match=r"signature"):
        load(p)


def test_load_legacy_xls_content_in_xlsx_raises(tmp_path):
    """A legacy .xls (OLE2) file renamed to .xlsx should be rejected."""
    p = tmp_path / "legacy.xlsx"
    p.write_bytes(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1" + b"\x00" * 100)
    with pytest.raises(ValueError, match=r"signature"):
        load(p)


def test_load_csv(tmp_path):
    """CSV files should be loadable through the pipeline."""
    p = tmp_path / "creds.csv"
    p.write_text("Password,s3cret!!\nToken,abc123")
    doc = load(p)
    assert isinstance(doc, GridDoc)
    assert doc.sheet_count == 1
    assert doc.cell_count == 4


def test_load_tsv(tmp_path):
    """TSV files should be loadable through the pipeline."""
    p = tmp_path / "data.tsv"
    p.write_text("Password\ts3cret!!\nToken\tabc123")
    doc = load(p)
    assert isinstance(doc, GridDoc)
    assert doc.sheet_count == 1
    assert doc.cell_count == 4


def test_load_xlsm(tmp_xlsx):
    """xlsm files should be loadable (openpyxl supports them natively)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Password"
    ws["B1"] = "s3cret!!"
    p = tmp_xlsx(wb, name="test.xlsm")
    wb.close()
    doc = load(p)
    assert isinstance(doc, GridDoc)
    assert doc.cell_count == 2


# ---------- GridDoc.relationships ----------


def test_relationships_returns_list(credential_xlsx):
    doc = load(credential_xlsx)
    rels = doc.relationships()
    assert isinstance(rels, list)


def test_relationships_finds_password(credential_xlsx):
    doc = load(credential_xlsx)
    rels = doc.relationships()
    assert len(rels) == 1
    assert rels[0].key == "Password"
    assert rels[0].value == "s3cret!!"


def test_relationships_returns_copy(credential_xlsx):
    doc = load(credential_xlsx)
    rels1 = doc.relationships()
    rels2 = doc.relationships()
    assert rels1 is not rels2


# ---------- GridDoc.credentials ----------


def test_credentials_returns_all_by_default(credential_xlsx):
    doc = load(credential_xlsx)
    creds = doc.credentials()
    assert len(creds) == len(doc.relationships())


def test_credentials_filters_by_min_confidence(credential_xlsx):
    doc = load(credential_xlsx)
    # The credential should have confidence > 0
    all_creds = doc.credentials(min_confidence=0.0)
    assert len(all_creds) >= 1

    # Filter with very high threshold should return nothing
    high_creds = doc.credentials(min_confidence=99999.0)
    assert len(high_creds) == 0


# ---------- Relationship ----------


def test_relationship_fields(credential_xlsx):
    doc = load(credential_xlsx)
    rel = doc.relationships()[0]
    assert isinstance(rel.key, str)
    assert isinstance(rel.value, str)
    assert isinstance(rel.confidence, float)
    assert isinstance(rel.reason, str)
    assert isinstance(rel.header_cell_id, int)
    assert isinstance(rel.value_cell_id, int)


def test_relationship_is_frozen(credential_xlsx):
    doc = load(credential_xlsx)
    rel = doc.relationships()[0]
    with pytest.raises(AttributeError):
        rel.key = "nope"


def test_relationship_repr(credential_xlsx):
    doc = load(credential_xlsx)
    rel = doc.relationships()[0]
    r = repr(rel)
    assert "Password" in r
    assert "s3cret!!" in r
    assert str(rel.confidence) in r


# ---------- gridmap.__init__ exports ----------


def test_exports_load():
    assert hasattr(gridmap, "load")
    assert gridmap.load is load


def test_exports_griddoc():
    assert hasattr(gridmap, "GridDoc")
    assert gridmap.GridDoc is GridDoc


def test_exports_relationship():
    assert hasattr(gridmap, "Relationship")
    assert gridmap.Relationship is Relationship


def test_exports_version():
    assert hasattr(gridmap, "version")
    v = gridmap.version()
    assert isinstance(v, str)
    assert len(v) > 0


# ---------- no credentials ----------


def test_no_credentials(tmp_xlsx):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Name"
    ws["B1"] = "Alice"
    ws["A2"] = "Age"
    ws["B2"] = "30"
    p = tmp_xlsx(wb)
    wb.close()

    doc = load(p)
    assert doc.relationships() == []
    assert doc.credentials() == []
