"""Tests for gridmap.extract — xlsx cell extraction via openpyxl."""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pytest

from gridmap.extract import (
    clean_comment,
    collect_merge_origins,
    extract_single_sheet,
    extract_workbook,
)


@pytest.fixture
def tmp_xlsx(tmp_path: Path):
    """Helper that writes an openpyxl Workbook to a temp file and returns the path."""

    def _write(wb: openpyxl.Workbook) -> Path:
        p = tmp_path / "test.xlsx"
        wb.save(p)
        return p

    return _write


# ---------- clean_comment ----------


def test_clean_comment_strips():
    assert clean_comment("  hello  ") == "hello"


def test_clean_comment_removes_prefix():
    assert clean_comment("Comment:\nhello world") == "hello world"


def test_clean_comment_no_prefix():
    assert clean_comment("just a note") == "just a note"


def test_clean_comment_empty():
    assert clean_comment("") == ""


# ---------- collect_merge_origins ----------


def test_collect_merge_origins():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.merge_cells("A1:B2")
    ws.merge_cells("D5:F7")
    origins = collect_merge_origins(ws)
    assert origins == {(1, 1), (5, 4)}
    wb.close()


# ---------- extract_single_sheet ----------


def test_extract_three_cells():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Name"
    ws["B1"] = "Alice"
    ws["A2"] = "Age"
    cells = extract_single_sheet(ws, "Sheet1", "visible")
    assert len(cells) == 3
    # Check first cell
    assert cells[0][0] == 1  # row
    assert cells[0][1] == 1  # col
    assert cells[0][2] == "Name"  # value
    assert cells[0][5] == "Sheet1"  # sheet_name
    wb.close()


def test_cell_with_comment():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Password"
    ws["A1"].comment = openpyxl.comments.Comment("secret note", "author")
    cells = extract_single_sheet(ws, "Sheet1", "visible")
    assert len(cells) == 1
    assert cells[0][4] == "secret note"  # comment field
    wb.close()


def test_cell_with_formula():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "=SUM(A1:A5)"
    cells = extract_single_sheet(ws, "Sheet1", "visible")
    assert len(cells) == 1
    assert cells[0][2] == "=SUM(A1:A5)"  # value
    assert cells[0][3] == "=SUM(A1:A5)"  # formula
    wb.close()


def test_hidden_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "data"
    cells = extract_single_sheet(ws, "SecretSheet", "hidden")
    assert cells[0][5] == "SecretSheet[HIDDEN]"
    wb.close()


def test_very_hidden_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "data"
    cells = extract_single_sheet(ws, "VerySecret", "veryHidden")
    assert cells[0][5] == "VerySecret[HIDDEN]"
    wb.close()


def test_merged_cell_origin():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Merged"
    ws["B1"] = "NotMerged"
    ws.merge_cells("A1:A3")
    cells = extract_single_sheet(ws, "Sheet1", "visible")
    merged_cell = [c for c in cells if c[0] == 1 and c[1] == 1][0]
    not_merged = [c for c in cells if c[0] == 1 and c[1] == 2][0]
    assert merged_cell[6] is True  # is_merged_origin
    assert not_merged[6] is False
    wb.close()


def test_empty_cells_skipped():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "data"
    # B1, C1 are empty (None)
    ws["D1"] = "more"
    cells = extract_single_sheet(ws, "Sheet1", "visible")
    assert len(cells) == 2
    wb.close()


def test_comment_prefix_stripped():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "value"
    ws["A1"].comment = openpyxl.comments.Comment("Comment:\nactual text", "author")
    cells = extract_single_sheet(ws, "Sheet1", "visible")
    assert cells[0][4] == "actual text"
    wb.close()


# ---------- extract_workbook ----------


def test_extract_workbook_one_sheet(tmp_xlsx):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Password"
    ws["B1"] = "s3cret!!"
    ws["A2"] = "Token"
    p = tmp_xlsx(wb)
    wb.close()

    sheets = extract_workbook(p)
    assert len(sheets) == 1
    assert len(sheets[0]) == 3


def test_extract_workbook_multiple_sheets(tmp_xlsx):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "data1"

    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "data2"
    ws2["B1"] = "data3"

    p = tmp_xlsx(wb)
    wb.close()

    sheets = extract_workbook(p)
    assert len(sheets) == 2
    assert len(sheets[0]) == 1
    assert len(sheets[1]) == 2


def test_extract_workbook_hidden_sheet(tmp_xlsx):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Visible"
    ws1["A1"] = "open"

    ws2 = wb.create_sheet("Hidden")
    ws2["A1"] = "secret"
    ws2.sheet_state = "hidden"

    p = tmp_xlsx(wb)
    wb.close()

    sheets = extract_workbook(p)
    assert len(sheets) == 2
    # Hidden sheet should have [HIDDEN] suffix
    hidden_cells = sheets[1]
    assert hidden_cells[0][5] == "Hidden[HIDDEN]"


def test_extract_workbook_accepts_path_object(tmp_xlsx):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "test"
    p = tmp_xlsx(wb)
    wb.close()

    # Should accept both str and Path
    sheets_str = extract_workbook(str(p))
    sheets_path = extract_workbook(p)
    assert len(sheets_str) == len(sheets_path)
