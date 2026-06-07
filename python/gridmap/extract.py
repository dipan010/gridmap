"""Extract raw cell tuples from xlsx files via openpyxl.

FIX 4: workbook is opened exactly once with data_only=False.
FIX 3: duplicate (row, col) entries merge comments into the existing cell.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl


def clean_comment(raw_text: str) -> str:
    """Strip whitespace and remove the 'Comment:\\n' prefix openpyxl adds."""
    text = raw_text.strip()
    if text.startswith("Comment:\n"):
        text = text[len("Comment:\n"):]
    return text.strip()


def collect_merge_origins(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> set[tuple[int, int]]:
    """Return the set of (min_row, min_col) for every merged cell range."""
    origins: set[tuple[int, int]] = set()
    for merge_range in worksheet.merged_cells.ranges:
        origins.add((merge_range.min_row, merge_range.min_col))
    return origins


def extract_single_sheet(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    sheet_name: str,
    sheet_state: str,
) -> list[tuple]:
    """Extract one worksheet into a list of raw cell tuples.

    Each tuple: (row, col, value, formula, comment, sheet_name, is_merged_origin)

    Hidden or veryHidden sheets get '[HIDDEN]' appended to sheet_name.
    Empty cells (value is None and no formula and no comment) are skipped.
    Duplicate (row, col) entries merge the comment into the first seen cell.
    """
    if sheet_state in ("hidden", "veryHidden"):
        sheet_name = f"{sheet_name}[HIDDEN]"

    merge_origins = collect_merge_origins(worksheet)

    # FIX 3: track seen coordinates for comment merging
    seen: dict[tuple[int, int], int] = {}
    cells: list[tuple] = []

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.row is None or cell.column is None:
                continue

            r = cell.row
            c = cell.column

            # Extract value and formula
            raw_value = cell.value
            formula = ""
            if isinstance(raw_value, str) and raw_value.startswith("="):
                formula = raw_value
                value = raw_value
            elif raw_value is not None:
                value = str(raw_value)
            else:
                value = ""

            # Extract comment
            comment = ""
            if cell.comment is not None:
                comment = clean_comment(cell.comment.text)

            # Skip truly empty cells
            if not value and not formula and not comment:
                continue

            coord = (r, c)
            is_merged_origin = coord in merge_origins

            # FIX 3: duplicate (row, col) — merge comment into existing
            if coord in seen:
                idx = seen[coord]
                if comment:
                    existing = cells[idx]
                    existing_comment = existing[4]
                    merged_comment = f"{existing_comment}\n{comment}".strip()
                    cells[idx] = (
                        existing[0],
                        existing[1],
                        existing[2],
                        existing[3],
                        merged_comment,
                        existing[5],
                        existing[6],
                    )
                continue

            seen[coord] = len(cells)
            cells.append((r, c, value, formula, comment, sheet_name, is_merged_origin))

    return cells


def extract_workbook(filepath: str | Path) -> list[list[tuple]]:
    """Open an xlsx workbook and extract all sheets as raw cell tuples.

    FIX 4: workbook is opened exactly once with data_only=False.

    Args:
        filepath: Path to the xlsx file.

    Returns:
        A list of lists — one inner list of tuples per sheet,
        matching the process_workbook input format.
    """
    filepath = Path(filepath)
    wb = openpyxl.load_workbook(filepath, data_only=False)

    sheets: list[list[tuple]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheet_state = ws.sheet_state
        cells = extract_single_sheet(ws, name, sheet_state)
        sheets.append(cells)

    wb.close()
    return sheets
